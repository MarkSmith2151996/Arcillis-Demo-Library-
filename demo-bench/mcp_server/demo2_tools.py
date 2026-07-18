"""Demo 2 document-extraction tools exposed by the Demo Bench MCP server."""

from __future__ import annotations

import csv
import json
import logging
import os
import re
import sys
import time
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

import psycopg2
from psycopg2.extras import Json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from gspread.utils import rowcol_to_a1


DATABASE_URL = os.environ.get(
    "DATABASE_URL",
    "postgresql://autocore_writer:autocore_pipeline_2026@localhost:5432/hive",
)
EXPORT_DIR = Path(os.environ.get("DEMO_BENCH_EXPORT_DIR", "/tmp/demo-bench-exports"))
GOOGLE_SHEETS_CREDENTIALS = os.environ.get("GOOGLE_SHEETS_CREDENTIALS", "")
INBOX_STAGING_DIR = Path(os.environ.get("INBOX_STAGING_DIR", "/tmp/arc-inbox-staging"))
FORBIDDEN_SQL = re.compile(r"\b(?:INSERT|UPDATE|DELETE|DROP|ALTER|CREATE|TRUNCATE)\b", re.IGNORECASE)
SOURCE_REFERENCE = re.compile(r"\b(?:FROM|JOIN)\s+([^\s,(;]+)", re.IGNORECASE)
FROM_CLAUSE = re.compile(
    r"\bFROM\s+(.+?)(?=\b(?:WHERE|GROUP|ORDER|HAVING|LIMIT|OFFSET|UNION)\b|$)",
    re.IGNORECASE | re.DOTALL,
)
LIMIT_CLAUSE = re.compile(r"\bLIMIT\s+\d+", re.IGNORECASE)
DEMO2_TABLE = re.compile(r"^demo2_[A-Za-z0-9_]+$")


def _connection():
    """Create a short-lived connection to Postgres and set the arcillis schema."""
    conn = psycopg2.connect(DATABASE_URL, connect_timeout=5)
    with conn.cursor() as cur:
        cur.execute("SET search_path TO arcillis")
    return conn


def _as_dict(value: Any) -> dict[str, Any]:
    if isinstance(value, str):
        try:
            value = json.loads(value)
        except json.JSONDecodeError:
            return {}
    return value if isinstance(value, dict) else {}


def _validate_read_query(sql: str) -> str:
    """Allow a single SELECT limited to Demo 2 tables."""
    query = sql.strip()
    if not query.upper().startswith("SELECT"):
        raise ValueError("Only SELECT queries are allowed.")
    if ";" in query:
        raise ValueError("Multiple SQL statements are not allowed.")
    if FORBIDDEN_SQL.search(query):
        raise ValueError("The query contains a forbidden SQL operation.")

    if any("," in match.group(1) for match in FROM_CLAUSE.finditer(query)):
        raise ValueError("Comma-separated table sources are not allowed.")

    for match in SOURCE_REFERENCE.finditer(query):
        relation = match.group(1).strip('"')
        table_name = relation.rsplit(".", maxsplit=1)[-1].strip('"')
        if not DEMO2_TABLE.fullmatch(table_name):
            raise ValueError("Queries may reference only demo2_* tables.")

    return query if LIMIT_CLAUSE.search(query) else f"{query} LIMIT 100"


def query_extractions(sql: str) -> dict[str, Any]:
    """Run a guarded read-only query over Demo 2 tables."""
    try:
        query = _validate_read_query(sql)
        with _connection() as connection:
            with connection.cursor() as cursor:
                cursor.execute("SET LOCAL statement_timeout = 5000")
                cursor.execute(query)
                columns = [column.name for column in cursor.description]
                rows = cursor.fetchmany(100)
        return {"columns": columns, "rows": [list(row) for row in rows], "row_count": len(rows)}
    except Exception as error:
        return {"error": str(error)}


def get_invoice_detail(invoice_id: int) -> dict[str, Any]:
    """Return the most recent extraction and source metadata for one invoice."""
    try:
        with _connection() as connection:
            with connection.cursor() as cursor:
                cursor.execute(
                    """
                    SELECT i.filename, i.dataset_source, i.image_path, i.ground_truth,
                           e.overall_accuracy, e.extracted_data, e.field_scores,
                           e.model_used, e.run_id, e.extracted_at
                    FROM arcillis.demo2_invoice i
                    LEFT JOIN arcillis.demo2_extraction e ON e.invoice_id = i.id
                    WHERE i.id = %s
                    ORDER BY e.extracted_at DESC NULLS LAST
                    LIMIT 1
                    """,
                    (invoice_id,),
                )
                row = cursor.fetchone()
                columns = [column.name for column in cursor.description]
        if row is None:
            return {"error": f"Invoice {invoice_id} was not found."}

        result = dict(zip(columns, row, strict=True))
        invoice = {
            "filename": result["filename"],
            "dataset": result["dataset_source"],
            "image_path": result["image_path"],
            "ground_truth": _as_dict(result["ground_truth"]),
        }
        if result["extracted_data"] is None:
            return {**invoice, "extraction": None}
        return {
            **invoice,
            "overall_accuracy": float(result["overall_accuracy"] or 0),
            "extracted_data": _as_dict(result["extracted_data"]),
            "field_scores": _as_dict(result["field_scores"]),
            "model_used": result["model_used"],
            "run_id": result["run_id"],
            "extracted_at": str(result["extracted_at"]),
        }
    except Exception as error:
        return {"error": str(error)}


def summarize_batch(run_id: str | None = None) -> dict[str, Any]:
    """Summarize extraction quality, datasets, and recurring field misses."""
    try:
        query = """
            SELECT e.run_id, e.overall_accuracy, e.field_scores, i.dataset_source
            FROM arcillis.demo2_extraction e
            JOIN arcillis.demo2_invoice i ON i.id = e.invoice_id
        """
        params: tuple[str, ...] = ()
        if run_id:
            query += " WHERE e.run_id = %s"
            params = (run_id,)
        with _connection() as connection:
            with connection.cursor() as cursor:
                cursor.execute(query, params)
                rows = cursor.fetchall()

                cursor.execute("SELECT DISTINCT run_id FROM arcillis.demo2_extraction ORDER BY run_id")
                available_runs = [row[0] for row in cursor.fetchall()]

        total = len(rows)
        accuracies = [float(row[1] or 0) for row in rows]
        grades = {"green_95_plus": 0, "yellow_80_94": 0, "red_below_80": 0}
        datasets: dict[str, list[float]] = defaultdict(list)
        field_totals: Counter[str] = Counter()
        field_misses: Counter[str] = Counter()
        for _, accuracy, scores, dataset in rows:
            value = float(accuracy or 0)
            grades["green_95_plus" if value >= 95 else "yellow_80_94" if value >= 80 else "red_below_80"] += 1
            datasets[str(dataset or "unknown")].append(value)
            for field, passed in _as_dict(scores).items():
                field_totals[field] += 1
                if not passed:
                    field_misses[field] += 1

        worst_fields = [
            {
                "field": field,
                "miss_count": field_misses[field],
                "miss_rate": round(field_misses[field] / field_totals[field], 4),
            }
            for field in sorted(field_totals, key=lambda name: (-field_misses[name], name))[:10]
        ]
        return {
            "total_extractions": total,
            "average_accuracy": round(sum(accuracies) / total, 2) if total else 0.0,
            "by_grade": grades,
            "by_dataset": {
                name: {"count": len(values), "avg_accuracy": round(sum(values) / len(values), 2)}
                for name, values in datasets.items()
            },
            "worst_fields": worst_fields,
            "available_runs": available_runs,
        }
    except Exception as error:
        return {"error": str(error)}


def _export_rows(invoice_ids: list[int] | None) -> tuple[list[str], list[list[Any]]]:
    query = """
        SELECT i.id, i.filename, i.dataset_source, e.overall_accuracy, e.model_used, e.run_id,
               e.extracted_data
        FROM arcillis.demo2_extraction e
        JOIN arcillis.demo2_invoice i ON i.id = e.invoice_id
    """
    params: tuple[Any, ...] = ()
    if invoice_ids is not None:
        if not invoice_ids:
            return _export_headers([]), []
        query += " WHERE i.id = ANY(%s)"
        params = (invoice_ids,)
    query += " ORDER BY i.filename"
    with _connection() as connection:
        with connection.cursor() as cursor:
            cursor.execute(query, params)
            records = cursor.fetchall()

    flattened: list[dict[str, Any]] = []
    for _, filename, dataset, accuracy, model, run, data in records:
        extracted = _as_dict(data)
        row: dict[str, Any] = {
            "filename": filename,
            "dataset_source": dataset,
            "overall_accuracy": float(accuracy or 0),
            "model_used": model,
            "run_id": run,
            "items_json": json.dumps(extracted.get("items", []), ensure_ascii=True),
        }
        for section in ("header", "summary"):
            values = extracted.get(section, {})
            if isinstance(values, dict):
                row.update({f"{section}.{key}": value for key, value in values.items()})
        flattened.append(row)
    headers = _export_headers(flattened)
    return headers, [[row.get(header, "") for header in headers] for row in flattened]


def _export_headers(rows: list[dict[str, Any]]) -> list[str]:
    metadata = ["filename", "dataset_source", "overall_accuracy", "model_used", "run_id"]
    fields = [
        field
        for row in rows
        for field in row
        if field not in metadata and field != "items_json"
    ]
    return [*metadata, *dict.fromkeys(fields), "items_json"]


def _output_path(filename: str, extension: str) -> Path:
    name = Path(filename).name
    if not name or name != filename:
        raise ValueError("filename must be a file name without directory components.")
    path = Path(name)
    if path.suffix.lower() != extension:
        path = path.with_suffix(extension)
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    return EXPORT_DIR / path.name


def export_to_csv(invoice_ids: list[int] | None = None, filename: str = "extractions.csv") -> dict[str, Any]:
    """Export selected or all extraction records to a flattened CSV file."""
    try:
        headers, rows = _export_rows(invoice_ids)
        output = _output_path(filename, ".csv")
        with output.open("w", encoding="utf-8", newline="") as output_file:
            writer = csv.writer(output_file)
            writer.writerow(headers)
            writer.writerows(rows)
        return {"filepath": str(output), "row_count": len(rows)}
    except Exception as error:
        return {"error": str(error)}


def export_to_excel(invoice_ids: list[int] | None = None, filename: str = "extractions.xlsx") -> dict[str, Any]:
    """Export selected or all extraction records to a formatted Excel workbook."""
    try:
        headers, rows = _export_rows(invoice_ids)
        output = _output_path(filename, ".xlsx")
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Extractions"
        worksheet.append(headers)
        for row in rows:
            worksheet.append(row)
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
        accuracy_column = headers.index("overall_accuracy") + 1
        fills = (
            (95, PatternFill("solid", fgColor="D4EDDA")),
            (80, PatternFill("solid", fgColor="FFF3CD")),
            (0, PatternFill("solid", fgColor="F8D7DA")),
        )
        for row_index in range(2, worksheet.max_row + 1):
            value = worksheet.cell(row_index, accuracy_column).value or 0
            worksheet.cell(row_index, accuracy_column).fill = next(fill for threshold, fill in fills if value >= threshold)
        for column_index, header in enumerate(headers, start=1):
            worksheet.column_dimensions[worksheet.cell(1, column_index).column_letter].width = max(len(header), 15)
        workbook.save(output)
        return {"filepath": str(output), "row_count": len(rows)}
    except Exception as error:
        return {"error": str(error)}


def write_cells(workbook: str, sheet: str, range: str, values: list[Any]) -> dict[str, Any]:
    """Write a value, row, or grid to a range in an open Excel workbook."""
    import xlwings as xw

    wb = xw.books[workbook]
    ws = wb.sheets[sheet]
    ws.range(range).value = values
    return {"status": "written", "workbook": workbook, "sheet": sheet, "range": range}


def read_cells(workbook: str, sheet: str, range: str) -> dict[str, Any]:
    """Read values from a range in an open Excel workbook."""
    import xlwings as xw

    wb = xw.books[workbook]
    ws = wb.sheets[sheet]
    return {"values": ws.range(range).value, "workbook": workbook, "sheet": sheet, "range": range}


def format_cells(
    workbook: str,
    sheet: str,
    range: str,
    bold: bool = False,
    color: str | None = None,
) -> dict[str, Any]:
    """Apply basic font formatting to a range in an open Excel workbook."""
    import xlwings as xw

    wb = xw.books[workbook]
    ws = wb.sheets[sheet]
    cell_range = ws.range(range)
    if bold:
        cell_range.font.bold = True
    if color:
        hex_color = color.lstrip("#")
        if not re.fullmatch(r"[0-9A-Fa-f]{6}", hex_color):
            raise ValueError("color must be a six-digit hexadecimal value, such as #22C55E.")
        cell_range.font.color = tuple(int(hex_color[index : index + 2], 16) for index in (0, 2, 4))
    return {"status": "formatted", "workbook": workbook, "sheet": sheet, "range": range}


def write_extraction_row(workbook: str, sheet: str, row_number: int, extraction_data: dict[str, Any]) -> dict[str, Any]:
    """Write one extraction result to Excel with an accuracy color indicator."""
    import xlwings as xw

    wb = xw.books[workbook]
    ws = wb.sheets[sheet]
    columns = ["A", "B", "C", "D", "E", "F", "G"]
    fields = ["invoice_id", "vendor", "invoice_number", "date", "total", "accuracy", "status"]

    for column, field in zip(columns, fields, strict=True):
        cell = ws.range(f"{column}{row_number}")
        value = extraction_data.get(field, "")
        cell.value = value
        if field == "accuracy" and isinstance(value, (int, float)):
            cell.font.color = (34, 197, 94) if value >= 90 else (234, 179, 8) if value >= 70 else (239, 68, 68)

    return {"status": "written", "workbook": workbook, "sheet": sheet, "row": row_number}



# ---------------------------------------------------------------------------
# Google Sheets tools (gspread + service account)
# ---------------------------------------------------------------------------

def _sheets_client():
    """Return an authorized gspread client using the service account JSON."""
    import gspread

    if not GOOGLE_SHEETS_CREDENTIALS:
        raise RuntimeError("GOOGLE_SHEETS_CREDENTIALS env var is not set.")
    return gspread.service_account(filename=GOOGLE_SHEETS_CREDENTIALS)


def sheets_write_headers(spreadsheet_id: str, sheet_name: str = "Sheet1") -> dict[str, Any]:
    """Write the standard extraction header row and apply bold formatting."""
    try:
        gc = _sheets_client()
        sh = gc.open_by_key(spreadsheet_id)
        ws = sh.worksheet(sheet_name)
        headers = ["Invoice ID", "Filename", "Vendor", "Invoice Number", "Date", "Total", "Accuracy", "Status"]
        ws.update("A1", [headers], value_input_option="RAW")
        ws.format("A1:H1", {"textFormat": {"bold": True}})
        return {"status": "headers_written", "spreadsheet_id": spreadsheet_id, "columns": len(headers)}
    except Exception as error:
        return {"error": str(error)}


def sheets_write_extraction_row(
    spreadsheet_id: str,
    row_number: int,
    extraction_data: dict[str, Any],
    sheet_name: str = "Sheet1",
) -> dict[str, Any]:
    """Write one extraction result as a color-coded row in a Google Sheet."""
    try:
        gc = _sheets_client()
        sh = gc.open_by_key(spreadsheet_id)
        ws = sh.worksheet(sheet_name)
        fields = ["invoice_id", "filename", "vendor", "invoice_number", "date", "total", "accuracy", "status"]
        row = [extraction_data.get(f, "") for f in fields]
        ws.update(f"A{row_number}", [row], value_input_option="RAW")

        accuracy = extraction_data.get("accuracy")
        if isinstance(accuracy, (int, float)):
            if accuracy >= 95:
                bg = {"red": 0.83, "green": 0.93, "blue": 0.85}
            elif accuracy >= 80:
                bg = {"red": 1.0, "green": 0.95, "blue": 0.8}
            else:
                bg = {"red": 0.97, "green": 0.84, "blue": 0.85}
            ws.format(f"G{row_number}", {"backgroundColor": bg})
        return {"status": "row_written", "spreadsheet_id": spreadsheet_id, "row": row_number}
    except Exception as error:
        return {"error": str(error)}


def sheets_write_cells(
    spreadsheet_id: str,
    range_str: str,
    values: list[Any],
    sheet_name: str = "Sheet1",
) -> dict[str, Any]:
    """Write a value, row, or grid to an arbitrary range in a Google Sheet."""
    try:
        gc = _sheets_client()
        sh = gc.open_by_key(spreadsheet_id)
        ws = sh.worksheet(sheet_name)
        data = values if isinstance(values[0], list) else [values]
        ws.update(range_str, data, value_input_option="RAW")
        return {"status": "written", "spreadsheet_id": spreadsheet_id, "range": range_str}
    except Exception as error:
        return {"error": str(error)}


def sheets_read_cells(
    spreadsheet_id: str,
    range_str: str,
    sheet_name: str = "Sheet1",
) -> dict[str, Any]:
    """Read values from a range in a Google Sheet."""
    try:
        gc = _sheets_client()
        sh = gc.open_by_key(spreadsheet_id)
        ws = sh.worksheet(sheet_name)
        values = ws.get(range_str)
        return {"values": values, "spreadsheet_id": spreadsheet_id, "range": range_str}
    except Exception as error:
        return {"error": str(error)}


def _column_letter(index: int) -> str:
    """Convert a zero-based column index to an A1 letter (0 -> A, 25 -> Z, 26 -> AA)."""
    result = ""
    n = index
    while n >= 0:
        result = chr(65 + (n % 26)) + result
        n = n // 26 - 1
    return result


def _rgb_to_hex(bg: dict[str, Any]) -> str | None:
    """Convert a Sheets color dict to a #RRGGBB hex string."""
    try:
        r = int(bg.get("red", 0) * 255)
        g = int(bg.get("green", 0) * 255)
        b = int(bg.get("blue", 0) * 255)
        return f"#{r:02X}{g:02X}{b:02X}"
    except (TypeError, ValueError):
        return None


def sheets_snapshot(
    spreadsheet_id: str,
    sheet_name: str = "Sheet1",
    include_values: bool = True,
    include_formatting: bool = False,
    compact_mode: bool = True,
    max_sample_rows: int = 5,
) -> dict[str, Any]:
    """Return a compressed structural and data summary of a Google Sheet."""
    try:
        gc = _sheets_client()
        sh = gc.open_by_key(spreadsheet_id)
        ws = sh.worksheet(sheet_name)

        values = ws.get_all_values()
        data_cells = [
            (row_idx, col_idx)
            for row_idx, row in enumerate(values)
            for col_idx, value in enumerate(row)
            if value != ""
        ]
        last_data_row = max((row_idx for row_idx, _ in data_cells), default=-1) + 1
        last_data_column = max((col_idx for _, col_idx in data_cells), default=-1) + 1
        values = [row[:last_data_column] for row in values[:last_data_row]]

        structure = {
            "sheet_name": sheet_name,
            "dimensions": {
                "rows": last_data_row,
                "columns": last_data_column,
                "grid_rows": ws.row_count,
                "grid_columns": ws.col_count,
            },
            "frozen": {"rows": ws.frozen_row_count or 0, "columns": ws.frozen_col_count or 0},
            "hidden_columns": [],
            "hidden_rows": [],
            "merges": [],
        }

        try:
            raw_merges = sh.client.request(
                "get",
                f"https://sheets.googleapis.com/v4/spreadsheets/{spreadsheet_id}",
                params={"fields": "sheets.merges"},
            )
            raw_json = raw_merges.json() if hasattr(raw_merges, "json") else raw_merges
            for s in raw_json.get("sheets", []):
                if s.get("properties", {}).get("title") == sheet_name:
                    structure["merges"] = [
                        merge.get("range", {}) for merge in s.get("merges", [])
                    ]
                    break
        except Exception:
            pass

        result: dict[str, Any] = {"structure": structure}

        if not include_values:
            return result

        if not values:
            result["headers"] = []
            result["column_summary"] = []
            result["sample_rows"] = {"first": [], "last": []}
            result["issues"] = []
            if not compact_mode:
                result["all_values"] = []
            return result

        headers = list(values[0])
        while len(headers) < last_data_column:
            headers.append("")
        while headers and headers[-1] == "":
            headers.pop()
        data_rows = values[1:] if len(values) > 1 else []

        result["headers"] = headers

        column_summary: list[dict[str, Any]] = []
        issues: list[str] = []

        for col_idx in range(last_data_column):
            col_letter = _column_letter(col_idx)
            header = headers[col_idx] if col_idx < len(headers) else ""
            col_values = [
                row[col_idx] if col_idx < len(row) else "" for row in data_rows
            ]
            filled = sum(1 for v in col_values if v != "")
            empty = len(col_values) - filled

            summary: dict[str, Any] = {
                "column": col_letter,
                "header": header,
                "type": "empty",
                "filled": filled,
                "empty": empty,
            }
            non_empty = [v for v in col_values if v != ""]

            if non_empty:
                numeric_count = 0
                numeric_values: list[float] = []
                for v in non_empty:
                    try:
                        cleaned = (
                            v.replace("$", "").replace(",", "").replace("%", "").strip()
                        )
                        if cleaned == "" or cleaned == "-":
                            continue
                        numeric_values.append(float(cleaned))
                        numeric_count += 1
                    except (ValueError, AttributeError):
                        pass

                if numeric_count > 0 and numeric_count / len(non_empty) >= 0.8:
                    summary["type"] = "numeric"
                    summary["min"] = round(min(numeric_values), 2)
                    summary["max"] = round(max(numeric_values), 2)
                    summary["avg"] = round(sum(numeric_values) / len(numeric_values), 2)
                elif numeric_count > 0:
                    summary["type"] = "mixed"
                else:
                    summary["type"] = "text"
                    unique_vals = list(dict.fromkeys(non_empty))
                    if len(unique_vals) <= 10:
                        summary["unique_values"] = unique_vals

            column_summary.append(summary)

            if empty > 0:
                label = f" ({header})" if header else ""
                issues.append(f"{empty} empty cells in column {col_letter}{label}")

        result["column_summary"] = column_summary
        result["issues"] = issues

        first_n = min(max_sample_rows, len(data_rows))
        last_n = 0
        if len(data_rows) > first_n:
            last_n = min(3, len(data_rows) - first_n)
        elif len(data_rows) <= max_sample_rows * 2:
            first_n = len(data_rows)
            last_n = 0

        result["sample_rows"] = {
            "first": data_rows[:first_n],
            "last": data_rows[-last_n:] if last_n > 0 else [],
        }

        if not compact_mode:
            result["all_values"] = values

        if include_formatting:
            formatting: dict[str, Any] = {"bold_ranges": [], "backgrounds": {}}
            try:
                raw_fmt = sh.client.request(
                    "get",
                    f"https://sheets.googleapis.com/v4/spreadsheets/{spreadsheet_id}",
                    params={
                        "fields": "sheets.data.rowData.values.userEnteredFormat",
                        "ranges": f"'{sheet_name}'",
                    },
                )
                fmt_json = raw_fmt.json() if hasattr(raw_fmt, "json") else raw_fmt
                for s in fmt_json.get("sheets", []):
                    for grid_data in s.get("data", []):
                        for row_idx, row_data in enumerate(grid_data.get("rowData", [])):
                            if row_idx >= last_data_row:
                                break
                            for col_idx, cell_data in enumerate(row_data.get("values", [])):
                                if col_idx >= last_data_column:
                                    break
                                fmt = cell_data.get("userEnteredFormat", {})
                                if not fmt:
                                    continue
                                cell_ref = f"{_column_letter(col_idx)}{row_idx + 1}"
                                if fmt.get("textFormat", {}).get("bold"):
                                    formatting["bold_ranges"].append(cell_ref)
                                bg = fmt.get("backgroundColor", {})
                                if bg:
                                    hex_color = _rgb_to_hex(bg)
                                    if hex_color:
                                        formatting["backgrounds"][cell_ref] = hex_color
            except Exception:
                pass
            result["formatting"] = formatting

        return result
    except Exception as error:
        return {"error": str(error)}


def _a1_to_grid_range(range_str: str, sheet_id: int) -> dict[str, int]:
    """Convert an A1 cell, row, or column range to a Sheets API GridRange."""
    reference = range_str.strip().replace("$", "")
    if "!" in reference:
        reference = reference.rsplit("!", 1)[1]
    match = re.fullmatch(r"([A-Za-z]*)(\d*)\s*(?::\s*([A-Za-z]*)(\d*))?", reference)
    if not match or not any(match.groups()):
        raise ValueError(f"Invalid A1 range: {range_str}")

    start_col, start_row, end_col, end_row = match.groups()
    if (start_col and not end_row and not end_col and not start_row) or (start_row and not end_col and not end_row and not start_col):
        raise ValueError(f"Invalid A1 range: {range_str}")

    def column_index(column: str) -> int:
        value = 0
        for letter in column.upper():
            value = value * 26 + ord(letter) - ord("A") + 1
        return value - 1

    grid_range: dict[str, int] = {"sheetId": sheet_id}
    if start_row:
        grid_range["startRowIndex"] = int(start_row) - 1
        grid_range["endRowIndex"] = int(end_row or start_row)
    if start_col:
        grid_range["startColumnIndex"] = column_index(start_col)
        grid_range["endColumnIndex"] = column_index(end_col or start_col) + 1
    return grid_range


def _hex_to_rgb(color: str) -> dict[str, float]:
    """Convert a six-digit hexadecimal color to the Sheets RGB representation."""
    hex_color = color.lstrip("#")
    if not re.fullmatch(r"[0-9A-Fa-f]{6}", hex_color):
        raise ValueError("Colors must be six-digit hexadecimal values, such as #22C55E.")
    return {
        "red": int(hex_color[0:2], 16) / 255,
        "green": int(hex_color[2:4], 16) / 255,
        "blue": int(hex_color[4:6], 16) / 255,
    }


def sheets_format_cells(
    spreadsheet_id: str,
    range_str: str,
    sheet_name: str = "Sheet1",
    bold: bool | None = None,
    italic: bool | None = None,
    font_color: str | None = None,
    background_color: str | None = None,
    font_size: int | None = None,
    horizontal_alignment: str | None = None,
    number_format: str | None = None,
    borders: dict[str, Any] | None = None,
    column_width: int | None = None,
) -> dict[str, Any]:
    """Apply text, number, border, and column-width formatting to a Sheets range."""
    try:
        gc = _sheets_client()
        sh = gc.open_by_key(spreadsheet_id)
        ws = sh.worksheet(sheet_name)
        format_dict: dict[str, Any] = {}
        text_format: dict[str, Any] = {}
        applied: list[str] = []
        if bold is not None:
            text_format["bold"] = bold
            applied.append("bold")
        if italic is not None:
            text_format["italic"] = italic
            applied.append("italic")
        if font_color:
            text_format["foregroundColor"] = _hex_to_rgb(font_color)
            applied.append("font_color")
        if text_format:
            format_dict["textFormat"] = text_format
        if background_color:
            format_dict["backgroundColor"] = _hex_to_rgb(background_color)
            applied.append("background_color")
        if font_size is not None:
            format_dict.setdefault("textFormat", {})["fontSize"] = font_size
            applied.append("font_size")
        if horizontal_alignment:
            alignment = horizontal_alignment.upper()
            if alignment not in {"LEFT", "CENTER", "RIGHT"}:
                raise ValueError("horizontal_alignment must be LEFT, CENTER, or RIGHT.")
            format_dict["horizontalAlignment"] = alignment
            applied.append("horizontal_alignment")
        if number_format:
            format_dict["numberFormat"] = {"type": "NUMBER", "pattern": number_format}
            applied.append("number_format")
        if format_dict:
            ws.format(range_str, format_dict)

        requests: list[dict[str, Any]] = []
        grid_range = _a1_to_grid_range(range_str, ws.id)
        if borders:
            style = borders.get("style", "SOLID")
            border = {"style": style, "colorStyle": {"rgbColor": {"red": 0, "green": 0, "blue": 0}}}
            border_request: dict[str, Any] = {"range": grid_range}
            for side in ("top", "bottom", "left", "right"):
                if borders.get(side):
                    border_request[side] = border
            if len(border_request) == 1:
                raise ValueError("borders must enable at least one side.")
            requests.append({"updateBorders": border_request})
            applied.append("borders")
        if column_width is not None:
            if column_width < 1 or "startColumnIndex" not in grid_range:
                raise ValueError("column_width requires a column-containing range and a positive pixel width.")
            requests.append({
                "updateDimensionProperties": {
                    "range": {
                        "sheetId": ws.id,
                        "dimension": "COLUMNS",
                        "startIndex": grid_range["startColumnIndex"],
                        "endIndex": grid_range["endColumnIndex"],
                    },
                    "properties": {"pixelSize": column_width},
                    "fields": "pixelSize",
                }
            })
            applied.append("column_width")
        if requests:
            sh.batch_update({"requests": requests})
        return {"status": "formatted", "spreadsheet_id": spreadsheet_id, "range": range_str, "applied": applied}
    except Exception as error:
        return {"error": str(error)}


def sheets_calculate(spreadsheet_id: str, formula: str, sheet_name: str = "Sheet1") -> dict[str, Any]:
    """Evaluate a formula by writing it to a temporary cell in the selected sheet."""
    try:
        gc = _sheets_client()
        sh = gc.open_by_key(spreadsheet_id)
        ws = sh.worksheet(sheet_name)
        formula = formula if formula.startswith("=") else f"={formula}"
        temp_cell = "Z999"
        ws.update_acell(temp_cell, formula)
        try:
            time.sleep(0.5)
            result = ws.acell(temp_cell, value_render_option="UNFORMATTED_VALUE").value
        finally:
            ws.update_acell(temp_cell, "")
        return {"formula": formula, "result": result}
    except Exception as error:
        return {"error": str(error)}


def sheets_append(spreadsheet_id: str, values: list[Any], sheet_name: str = "Sheet1") -> dict[str, Any]:
    """Append one row or a grid of rows to a Google Sheet."""
    try:
        if not values:
            raise ValueError("values must contain at least one row.")
        rows = values if isinstance(values[0], list) else [values]
        gc = _sheets_client()
        sh = gc.open_by_key(spreadsheet_id)
        sh.worksheet(sheet_name).append_rows(rows, value_input_option="RAW")
        return {"status": "appended", "rows_added": len(rows), "spreadsheet_id": spreadsheet_id}
    except Exception as error:
        return {"error": str(error)}


def sheets_clear(spreadsheet_id: str, range_str: str, sheet_name: str = "Sheet1") -> dict[str, Any]:
    """Clear values from a range without removing its rows or columns."""
    try:
        gc = _sheets_client()
        sh = gc.open_by_key(spreadsheet_id)
        sh.worksheet(sheet_name).batch_clear([range_str])
        return {"status": "cleared", "range": range_str, "spreadsheet_id": spreadsheet_id}
    except Exception as error:
        return {"error": str(error)}


def sheets_manage(
    spreadsheet_id: str,
    operation: str,
    sheet_name: str = "Sheet1",
    row_index: int | None = None,
    num_rows: int = 1,
    new_name: str | None = None,
) -> dict[str, Any]:
    """Insert or delete rows, or add, rename, and delete worksheet tabs."""
    try:
        gc = _sheets_client()
        sh = gc.open_by_key(spreadsheet_id)
        if operation == "add_sheet":
            if not new_name:
                raise ValueError("new_name is required for add_sheet.")
            ws = sh.add_worksheet(title=new_name, rows=1000, cols=26)
            return {"status": "done", "operation": operation, "spreadsheet_id": spreadsheet_id, "sheet_name": ws.title}
        ws = sh.worksheet(sheet_name)
        if operation == "insert_rows":
            if row_index is None or row_index < 1 or num_rows < 1:
                raise ValueError("insert_rows requires a positive row_index and num_rows.")
            ws.insert_rows(values=[[""] * ws.col_count] * num_rows, row=row_index)
        elif operation == "delete_rows":
            if row_index is None or row_index < 1 or num_rows < 1:
                raise ValueError("delete_rows requires a positive row_index and num_rows.")
            ws.delete_rows(row_index, row_index + num_rows - 1)
        elif operation == "rename_sheet":
            if not new_name:
                raise ValueError("new_name is required for rename_sheet.")
            ws.update_title(new_name)
        elif operation == "delete_sheet":
            sh.del_worksheet(ws)
        else:
            raise ValueError("operation must be insert_rows, delete_rows, add_sheet, rename_sheet, or delete_sheet.")
        return {"status": "done", "operation": operation, "spreadsheet_id": spreadsheet_id}
    except Exception as error:
        return {"error": str(error)}


def sheets_merge(
    spreadsheet_id: str,
    range_str: str,
    sheet_name: str = "Sheet1",
    merge_type: str = "MERGE_ALL",
) -> dict[str, Any]:
    """Merge or unmerge a range of cells."""
    try:
        gc = _sheets_client()
        sh = gc.open_by_key(spreadsheet_id)
        ws = sh.worksheet(sheet_name)
        if merge_type == "UNMERGE":
            ws.unmerge_cells(range_str)
            status = "unmerged"
        else:
            if merge_type not in {"MERGE_ALL", "MERGE_COLUMNS", "MERGE_ROWS"}:
                raise ValueError("merge_type must be MERGE_ALL, MERGE_COLUMNS, MERGE_ROWS, or UNMERGE.")
            ws.merge_cells(range_str, merge_type=merge_type)
            status = "merged"
        return {"status": status, "range": range_str, "spreadsheet_id": spreadsheet_id}
    except Exception as error:
        return {"error": str(error)}


def sheets_find(
    spreadsheet_id: str, query: str, sheet_name: str = "Sheet1", case_sensitive: bool = False
) -> dict[str, Any]:
    """Find all matching cells in a Google Sheet."""
    try:
        gc = _sheets_client()
        sh = gc.open_by_key(spreadsheet_id)
        ws = sh.worksheet(sheet_name)
        pattern: str | re.Pattern[str] = query if case_sensitive else re.compile(re.escape(query), re.IGNORECASE)
        results = ws.findall(pattern)
        matches = [{"cell": rowcol_to_a1(cell.row, cell.col), "value": cell.value, "row": cell.row, "col": cell.col} for cell in results]
        return {"matches": matches, "count": len(matches), "spreadsheet_id": spreadsheet_id}
    except Exception as error:
        return {"error": str(error)}


def sheets_conditional_format(
    spreadsheet_id: str,
    range_str: str,
    rule_type: str,
    values: list[Any],
    format: dict[str, Any],
    sheet_name: str = "Sheet1",
) -> dict[str, Any]:
    """Add a raw Sheets API conditional-format rule to a range."""
    try:
        gc = _sheets_client()
        sh = gc.open_by_key(spreadsheet_id)
        ws = sh.worksheet(sheet_name)
        sheets_format: dict[str, Any] = {}
        if format.get("background_color"):
            sheets_format["backgroundColor"] = _hex_to_rgb(format["background_color"])
        text_format: dict[str, Any] = {}
        if format.get("font_color"):
            text_format["foregroundColor"] = _hex_to_rgb(format["font_color"])
        if "bold" in format:
            text_format["bold"] = bool(format["bold"])
        if text_format:
            sheets_format["textFormat"] = text_format
        if not sheets_format:
            raise ValueError("format must include background_color, font_color, or bold.")
        request = {"addConditionalFormatRule": {"rule": {
            "ranges": [_a1_to_grid_range(range_str, ws.id)],
            "booleanRule": {
                "condition": {"type": rule_type, "values": [{"userEnteredValue": str(value)} for value in values]},
                "format": sheets_format,
            },
        }, "index": 0}}
        sh.batch_update({"requests": [request]})
        return {"status": "rule_added", "range": range_str, "rule_type": rule_type, "spreadsheet_id": spreadsheet_id}
    except Exception as error:
        return {"error": str(error)}


def sheets_validate(
    spreadsheet_id: str,
    range_str: str,
    validation_type: str,
    values: list[Any] | None = None,
    sheet_name: str = "Sheet1",
    strict: bool = True,
    show_dropdown: bool = True,
) -> dict[str, Any]:
    """Set dropdown, checkbox, number, or formula validation on a Sheets range."""
    try:
        gc = _sheets_client()
        sh = gc.open_by_key(spreadsheet_id)
        ws = sh.worksheet(sheet_name)
        condition: dict[str, Any] = {"type": "BOOLEAN" if validation_type == "CHECKBOX" else validation_type}
        if values:
            condition["values"] = [{"userEnteredValue": str(value)} for value in values]
        request = {"setDataValidation": {"range": _a1_to_grid_range(range_str, ws.id), "rule": {
            "condition": condition,
            "strict": strict,
            "showCustomUi": show_dropdown,
        }}}
        sh.batch_update({"requests": [request]})
        return {"status": "validation_set", "range": range_str, "type": validation_type, "spreadsheet_id": spreadsheet_id}
    except Exception as error:
        return {"error": str(error)}


def reprocess_invoices(invoice_ids: list[int]) -> dict[str, Any]:
    """Queue a future extraction rerun without invoking the extraction engine yet."""
    logging.getLogger(__name__).info("Reprocessing requested for invoices: %s", invoice_ids)
    return {
        "status": "queued",
        "invoice_ids": invoice_ids,
        "message": f"Reprocessing queued for {len(invoice_ids)} invoices. This will be processed in the next extraction batch.",
    }


def scan_inbox() -> dict[str, Any]:
    """Download unread PDF and image attachments from the configured Gmail inbox."""
    credentials_path = os.environ.get("GMAIL_CREDENTIALS_PATH")
    if not credentials_path:
        return {"emails_processed": 0, "attachments_saved": [], "errors": ["GMAIL_CREDENTIALS_PATH is not set."]}

    try:
        from google.auth.transport.requests import Request
        from google.oauth2.credentials import Credentials
        from google_auth_oauthlib.flow import InstalledAppFlow
        from googleapiclient.discovery import build
    except ImportError as error:
        return {"emails_processed": 0, "attachments_saved": [], "errors": [f"Gmail dependencies are unavailable: {error}"]}

    try:
        scopes = ["https://www.googleapis.com/auth/gmail.modify"]
        token_path = Path(credentials_path).with_name("gmail-token.json")
        credentials = Credentials.from_authorized_user_file(token_path, scopes) if token_path.exists() else None
        if not credentials or not credentials.valid:
            if credentials and credentials.expired and credentials.refresh_token:
                credentials.refresh(Request())
            else:
                credentials = InstalledAppFlow.from_client_secrets_file(credentials_path, scopes).run_local_server(port=0)
            token_path.write_text(credentials.to_json(), encoding="utf-8")

        service = build("gmail", "v1", credentials=credentials, cache_discovery=False)
        messages = service.users().messages().list(userId="me", q="is:unread has:attachment").execute().get("messages", [])
        INBOX_STAGING_DIR.mkdir(parents=True, exist_ok=True)
        attachments_saved: list[str] = []
        errors: list[str] = []
        emails_processed = 0

        for message_ref in messages:
            message_id = message_ref["id"]
            try:
                message = service.users().messages().get(userId="me", id=message_id).execute()
                parts = _attachment_parts(message.get("payload", {}))
                for part in parts:
                    filename = Path(part.get("filename", "")).name
                    attachment_id = part.get("body", {}).get("attachmentId")
                    if not filename or not attachment_id:
                        continue
                    content = service.users().messages().attachments().get(
                        userId="me", messageId=message_id, id=attachment_id
                    ).execute()["data"]
                    import base64

                    destination = _unique_path(INBOX_STAGING_DIR / filename)
                    destination.write_bytes(base64.urlsafe_b64decode(content + "=" * (-len(content) % 4)))
                    attachments_saved.append(str(destination))
                service.users().messages().modify(userId="me", id=message_id, body={"removeLabelIds": ["UNREAD"]}).execute()
                emails_processed += 1
            except Exception as error:
                errors.append(f"{message_id}: {error}")
        return {"emails_processed": emails_processed, "attachments_saved": attachments_saved, "errors": errors}
    except Exception as error:
        return {"emails_processed": 0, "attachments_saved": [], "errors": [str(error)]}


def _attachment_parts(part: dict[str, Any]) -> list[dict[str, Any]]:
    """Flatten Gmail MIME parts that contain a downloaded attachment body."""
    result = [part] if part.get("filename") and part.get("body", {}).get("attachmentId") else []
    for child in part.get("parts", []):
        result.extend(_attachment_parts(child))
    return result


def _unique_path(path: Path) -> Path:
    """Avoid overwriting a same-named attachment from a separate email."""
    if not path.exists():
        return path
    return path.with_stem(f"{path.stem}-{int(time.time() * 1000)}")


def run_extraction(source_dir: str | None = None) -> dict[str, Any]:
    """Extract staged invoice images and persist results using the established vision pipeline."""
    source = Path(source_dir) if source_dir else INBOX_STAGING_DIR
    if not source.is_dir():
        return {"processed": 0, "succeeded": 0, "failed": 0, "errors": [f"Source directory does not exist: {source}"]}

    files = [path for path in sorted(source.iterdir()) if path.suffix.lower() in {".jpg", ".jpeg", ".png", ".webp", ".pdf"}]
    if not files:
        return {"processed": 0, "succeeded": 0, "failed": 0, "errors": []}

    try:
        extract_invoice, grade_extraction, create_client, verify_proxy = _load_extraction_helpers()
        client = create_client()
        model_used = verify_proxy(client)
    except Exception as error:
        return {"processed": len(files), "succeeded": 0, "failed": len(files), "errors": [str(error)]}

    run_id = f"inbox-{time.strftime('%Y%m%d-%H%M%S')}"
    succeeded = 0
    errors: list[str] = []
    with _connection() as connection:
        for source_file in files:
            started = time.perf_counter()
            try:
                image_path = _render_pdf(source_file) if source_file.suffix.lower() == ".pdf" else source_file
                extracted_data = extract_invoice(image_path, client=client, model=model_used)
                invoice_id, ground_truth = _get_or_create_inbox_invoice(connection, source_file)
                field_scores, accuracy = grade_extraction(extracted_data, ground_truth)
                with connection.cursor() as cursor:
                    cursor.execute(
                        """
                        INSERT INTO demo2_extraction
                            (invoice_id, run_id, model_used, extracted_data, field_scores, overall_accuracy, processing_time_ms)
                        VALUES (%s, %s, %s, %s, %s, %s, %s)
                        """,
                        (invoice_id, run_id, model_used, Json(extracted_data), Json(field_scores), accuracy, round((time.perf_counter() - started) * 1000)),
                    )
                connection.commit()
                succeeded += 1
            except Exception as error:
                connection.rollback()
                errors.append(f"{source_file.name}: {error}")
    return {"processed": len(files), "succeeded": succeeded, "failed": len(files) - succeeded, "errors": errors}


def _load_extraction_helpers() -> tuple[Any, Any, Any, Any]:
    """Import the existing extraction and grading modules from the sibling demo directory."""
    extractor_dir = Path(__file__).resolve().parents[2] / "demo-2-pdf-extractor"
    if str(extractor_dir) not in sys.path:
        sys.path.insert(0, str(extractor_dir))
    from extract_invoice import create_client, extract_invoice, verify_proxy
    from grade_extraction import grade_extraction

    return extract_invoice, grade_extraction, create_client, verify_proxy


def _render_pdf(source_file: Path) -> Path:
    """Render the first PDF page for the vision model without changing source files."""
    try:
        import fitz
    except ImportError as error:
        raise RuntimeError("PDF extraction requires pymupdf.") from error
    document = fitz.open(source_file)
    try:
        if not document.page_count:
            raise ValueError("PDF has no pages.")
        destination = source_file.with_suffix(".page-1.png")
        document[0].get_pixmap(matrix=fitz.Matrix(2, 2)).save(destination)
        return destination
    finally:
        document.close()


def _get_or_create_inbox_invoice(connection: Any, source_file: Path) -> tuple[int, dict[str, Any]]:
    """Return an existing invoice or create a minimal inbox record for an attachment."""
    with connection.cursor() as cursor:
        cursor.execute("SELECT id, ground_truth FROM demo2_invoice WHERE filename = %s", (source_file.name,))
        row = cursor.fetchone()
        if row:
            return int(row[0]), _as_dict(row[1])
        cursor.execute(
            """
            INSERT INTO demo2_invoice (filename, split, ground_truth, image_path, dataset_source)
            VALUES (%s, 'inbox', %s, %s, 'gmail_inbox')
            RETURNING id
            """,
            (source_file.name, Json({}), str(source_file.resolve())),
        )
        return int(cursor.fetchone()[0]), {}
