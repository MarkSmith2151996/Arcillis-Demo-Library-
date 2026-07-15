"""CSV and Excel export dock for checked Demo Bench extraction results."""

from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any

from PySide6.QtWidgets import QComboBox, QFileDialog, QLabel, QMessageBox, QPushButton, QVBoxLayout, QWidget

from widgets.results_table_widget import ResultsTableWidget


class ExportWidget(QWidget):
    """Export selected Results Table rows, or all rows after confirmation."""

    def __init__(self, results_table: ResultsTableWidget, parent: QWidget | None = None) -> None:
        super().__init__(parent)
        self.results_table = results_table
        self.selection_label = QLabel("0 rows selected")
        self.format_box = QComboBox()
        self.format_box.addItems(["CSV", "Excel (.xlsx)"])
        self.export_button = QPushButton("Export")
        self.status = QLabel()
        self.status.setWordWrap(True)

        layout = QVBoxLayout(self)
        layout.addWidget(self.selection_label)
        layout.addWidget(self.format_box)
        layout.addWidget(self.export_button)
        layout.addWidget(self.status)
        layout.addStretch()

        self.export_button.clicked.connect(self.export)
        results_table.selection_changed.connect(self.update_selection_count)

    def update_selection_count(self, count: int) -> None:
        self.selection_label.setText(f"{count} row{'s' if count != 1 else ''} selected")

    def export(self) -> None:
        """Ask for a destination and write the checked, or confirmed all, rows."""
        rows = self.results_table.get_selected_data()
        if not rows:
            if QMessageBox.question(
                self,
                "Export all rows?",
                "No rows are checked. Export every result instead?",
            ) != QMessageBox.StandardButton.Yes:
                return
            rows = self.results_table.get_all_data()
        if not rows:
            self.status.setText("No extraction rows are available to export.")
            return

        excel = self.format_box.currentText().startswith("Excel")
        extension = "xlsx" if excel else "csv"
        path, _ = QFileDialog.getSaveFileName(self, "Export extraction results", f"extractions.{extension}", f"*.{extension}")
        if not path:
            return
        try:
            headers, values = self._table(rows)
            if excel:
                self._write_excel(Path(path), headers, values)
            else:
                self._write_csv(Path(path), headers, values)
            self.status.setText(f"Exported {len(values)} rows to {path}")
        except Exception as error:
            self.status.setText(f"Export failed: {error}")

    def _table(self, rows: list[dict[str, Any]]) -> tuple[list[str], list[dict[str, Any]]]:
        records: list[dict[str, Any]] = []
        fields: list[str] = []
        for source in rows:
            data = self._json(source["extracted_data"])
            record = {
                "filename": Path(str(source["filename"])).name,
                "dataset_source": source["dataset_source"] or "",
                "overall_accuracy": source["overall_accuracy"] or "",
                "model_used": source["model_used"] or "",
                "run_id": source["run_id"] or "",
                "items": json.dumps(data.get("items", []), ensure_ascii=True),
            }
            for section in ("header", "summary"):
                for key, value in self._flatten(data.get(section, {}), section).items():
                    record[key] = value
                    if key not in fields:
                        fields.append(key)
            records.append(record)
        headers = ["filename", "dataset_source", "overall_accuracy", "model_used", "run_id", *fields, "items"]
        return headers, records

    @staticmethod
    def _json(value: Any) -> dict[str, Any]:
        if isinstance(value, str):
            value = json.loads(value)
        return value if isinstance(value, dict) else {}

    @classmethod
    def _flatten(cls, value: Any, prefix: str) -> dict[str, Any]:
        if isinstance(value, dict):
            return {key: child for name, item in value.items() for key, child in cls._flatten(item, f"{prefix}.{name}").items()}
        return {prefix: value}

    @staticmethod
    def _write_csv(path: Path, headers: list[str], rows: list[dict[str, Any]]) -> None:
        with path.open("w", encoding="utf-8", newline="") as output:
            writer = csv.DictWriter(output, fieldnames=headers)
            writer.writeheader()
            writer.writerows(rows)

    @staticmethod
    def _write_excel(path: Path, headers: list[str], rows: list[dict[str, Any]]) -> None:
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
        except ImportError as error:
            raise RuntimeError("Excel export requires openpyxl. Install it in the Demo Bench venv.") from error
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Extractions"
        sheet.append(headers)
        for cell in sheet[1]:
            cell.font = Font(bold=True)
        grade_column = headers.index("overall_accuracy") + 1
        for record in rows:
            sheet.append([record.get(header, "") for header in headers])
            try:
                grade = float(record["overall_accuracy"])
            except (TypeError, ValueError):
                grade = 0.0
            color = "D4EDDA" if grade >= 95 else "FFF3CD" if grade >= 80 else "F8D7DA"
            sheet.cell(sheet.max_row, grade_column).fill = PatternFill("solid", fgColor=color)
        for cells in sheet.columns:
            sheet.column_dimensions[cells[0].column_letter].width = min(max(len(str(cell.value or "")) for cell in cells) + 2, 60)
        workbook.save(path)
