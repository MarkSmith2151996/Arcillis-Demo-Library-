"""Integration test for Demo 2 output writers."""

from __future__ import annotations

import csv
from pathlib import Path

from openpyxl import load_workbook

from server import SheetsCredentialsError, write_to_csv, write_to_excel, write_to_sheets


RUN_ID = "mychen76-test-run"
EXPECTED_DOCUMENTS = 125


def main() -> int:
    """Write and validate all configured output destinations."""
    output_dir = Path(__file__).resolve().parent / "output"
    excel_result = write_to_excel(run_id=RUN_ID, output_path=str(output_dir / f"{RUN_ID}.xlsx"))
    csv_result = write_to_csv(run_id=RUN_ID, output_path=str(output_dir / f"{RUN_ID}.csv"))

    workbook = load_workbook(excel_result["output_path"], read_only=True)
    excel_rows = workbook.active.max_row - 1
    with Path(csv_result["output_path"]).open(encoding="utf-8", newline="") as csv_file:
        csv_rows = sum(1 for _ in csv.reader(csv_file)) - 1

    if excel_rows != EXPECTED_DOCUMENTS:
        raise AssertionError(f"Excel contains {excel_rows} documents; expected {EXPECTED_DOCUMENTS}.")
    if csv_rows != EXPECTED_DOCUMENTS:
        raise AssertionError(f"CSV contains {csv_rows} documents; expected {EXPECTED_DOCUMENTS}.")
    if excel_rows != csv_rows:
        raise AssertionError("Excel and CSV row counts do not match.")

    print(f"Excel: {excel_result['output_path']} ({excel_rows} documents)")
    print(f"CSV: {csv_result['output_path']} ({csv_rows} documents)")
    try:
        sheets_result = write_to_sheets(run_id=RUN_ID)
        print(f"Sheets: {sheets_result['spreadsheet_url']} ({sheets_result['row_count']} documents)")
    except SheetsCredentialsError:
        print("Sheets test skipped — no service account configured")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
