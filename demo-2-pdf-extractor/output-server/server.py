"""MCP output tools for Demo 2 extraction results."""

from __future__ import annotations

import csv
import json
from pathlib import Path
from typing import Any

import psycopg2
from mcp.server.fastmcp import FastMCP
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill


SERVER_DIR = Path(__file__).resolve().parent
CONFIG_PATH = SERVER_DIR / "config.json"
mcp = FastMCP("Demo 2 Output Server")


class SheetsCredentialsError(RuntimeError):
    """Raised when Google Sheets credentials have not been configured."""


def load_config() -> dict[str, Any]:
    """Load the server's destination settings."""
    return json.loads(CONFIG_PATH.read_text(encoding="utf-8"))


def flatten_json(obj: Any, prefix: str = "") -> dict[str, Any]:
    """Flatten nested dictionaries and lists into dotted, indexed scalar paths."""
    if isinstance(obj, dict):
        flattened: dict[str, Any] = {}
        for key, value in obj.items():
            path = f"{prefix}.{key}" if prefix else str(key)
            flattened.update(flatten_json(value, path))
        return flattened

    if isinstance(obj, list):
        flattened = {}
        for index, value in enumerate(obj):
            path = f"{prefix}[{index}]" if prefix else f"[{index}]"
            flattened.update(flatten_json(value, path))
        return flattened

    return {prefix: obj}


def fetch_extractions(run_id: str) -> list[dict[str, Any]]:
    """Fetch ordered extraction results for one run directly from Postgres."""
    query = """
        SELECT e.invoice_id, e.extracted_data, e.overall_accuracy, e.field_scores, i.filename
        FROM arcillis.demo2_extraction e
        JOIN arcillis.demo2_invoice i ON e.invoice_id = i.id
        WHERE e.run_id = %s
        ORDER BY i.filename
    """
    with psycopg2.connect(load_config()["database_url"]) as connection:
        with connection.cursor() as cursor:
            cursor.execute(query, (run_id,))
            columns = [column.name for column in cursor.description]
            return [dict(zip(columns, row, strict=True)) for row in cursor.fetchall()]


def build_table(extractions: list[dict[str, Any]]) -> tuple[list[str], list[list[Any]]]:
    """Return consistent metadata-first columns and rows for every output type."""
    flattened_rows: list[dict[str, Any]] = []
    field_names: list[str] = []

    for index, extraction in enumerate(extractions, start=1):
        data = extraction.get("extracted_data", extraction)
        if not isinstance(data, dict):
            raise ValueError(f"Extraction {index} must be a JSON object.")

        row = {
            "filename": extraction.get("filename", f"document_{index}"),
            "overall_accuracy": extraction.get("overall_accuracy", ""),
            **flatten_json(data),
        }
        flattened_rows.append(row)
        for field_name in row:
            if field_name not in {"filename", "overall_accuracy"} and field_name not in field_names:
                field_names.append(field_name)

    headers = ["filename", "overall_accuracy", *field_names]
    return headers, [[row.get(header, "") for header in headers] for row in flattened_rows]


def resolve_extractions(
    run_id: str | None, extractions: list[dict[str, Any]] | None
) -> tuple[str, list[dict[str, Any]]]:
    """Load one requested source and provide a stable output file stem."""
    if bool(run_id) == bool(extractions):
        raise ValueError("Provide exactly one of run_id or extractions.")
    if run_id:
        return run_id, fetch_extractions(run_id)
    return "extractions", extractions or []


def default_output_path(stem: str, extension: str) -> Path:
    """Resolve the configured local output destination."""
    output_dir = Path(load_config().get("output_dir", "output"))
    if not output_dir.is_absolute():
        output_dir = SERVER_DIR / output_dir
    return output_dir / f"{stem}.{extension}"


def resolve_output_path(output_path: str | None, stem: str, extension: str) -> Path:
    """Use an explicit output path or the server's configured default."""
    return Path(output_path) if output_path else default_output_path(stem, extension)


def write_excel_file(headers: list[str], rows: list[list[Any]], output_path: Path) -> None:
    """Write a formatted workbook with one document per row."""
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Extractions"
    worksheet.append(headers)
    for row in rows:
        worksheet.append(row)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    alternate_fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in worksheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
    for row_index in range(2, worksheet.max_row + 1):
        if row_index % 2 == 0:
            for cell in worksheet[row_index]:
                cell.fill = alternate_fill
    for column_cells in worksheet.columns:
        letter = column_cells[0].column_letter
        worksheet.column_dimensions[letter].width = min(
            max(len(str(cell.value or "")) for cell in column_cells) + 2, 60
        )
    worksheet.freeze_panes = "A2"
    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)


def write_csv_file(headers: list[str], rows: list[list[Any]], output_path: Path) -> None:
    """Write the same table structure as a CSV fallback."""
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", encoding="utf-8", newline="") as output_file:
        writer = csv.writer(output_file)
        writer.writerow(headers)
        writer.writerows(rows)


def sheets_client() -> Any:
    """Return an authorized Sheets client or a setup-focused error."""
    credentials = load_config().get("google_service_account_credentials", "")
    credentials_path = Path(credentials).expanduser() if credentials else None
    if not credentials_path or not credentials_path.is_file():
        raise SheetsCredentialsError(
            "Google Sheets is not configured. Set google_service_account_credentials in "
            "config.json to a readable service-account JSON file, then share the target "
            "spreadsheet with that service account."
        )
    import gspread

    return gspread.service_account(filename=str(credentials_path))


@mcp.tool()
def write_to_excel(
    run_id: str | None = None,
    extractions: list[dict[str, Any]] | None = None,
    output_path: str | None = None,
) -> dict[str, Any]:
    """Write extraction JSON to a formatted Excel workbook."""
    stem, source = resolve_extractions(run_id, extractions)
    headers, rows = build_table(source)
    destination = resolve_output_path(output_path, stem, "xlsx")
    write_excel_file(headers, rows, destination)
    return {"output_path": str(destination), "row_count": len(rows), "columns": headers}


@mcp.tool()
def write_to_csv(
    run_id: str | None = None,
    extractions: list[dict[str, Any]] | None = None,
    output_path: str | None = None,
) -> dict[str, Any]:
    """Write extraction JSON to a CSV file."""
    stem, source = resolve_extractions(run_id, extractions)
    headers, rows = build_table(source)
    destination = resolve_output_path(output_path, stem, "csv")
    write_csv_file(headers, rows, destination)
    return {"output_path": str(destination), "row_count": len(rows), "columns": headers}


@mcp.tool()
def write_to_sheets(
    run_id: str | None = None,
    extractions: list[dict[str, Any]] | None = None,
    spreadsheet_id: str | None = None,
    sheet_name: str | None = None,
) -> dict[str, Any]:
    """Create or update a Google Sheets tab with flattened extraction data."""
    stem, source = resolve_extractions(run_id, extractions)
    headers, rows = build_table(source)
    client = sheets_client()
    spreadsheet = client.open_by_key(spreadsheet_id) if spreadsheet_id else client.create(f"Demo 2 {stem}")
    tab_name = sheet_name or stem
    try:
        worksheet = spreadsheet.worksheet(tab_name)
        worksheet.clear()
    except Exception as error:
        if error.__class__.__name__ != "WorksheetNotFound":
            raise
        worksheet = spreadsheet.add_worksheet(title=tab_name, rows=max(len(rows) + 1, 100), cols=max(len(headers), 26))
    worksheet.update("A1", [headers, *rows], value_input_option="USER_ENTERED")
    worksheet.freeze(rows=1)
    return {
        "spreadsheet_id": spreadsheet.id,
        "spreadsheet_url": spreadsheet.url,
        "sheet_name": tab_name,
        "row_count": len(rows),
        "columns": headers,
    }


if __name__ == "__main__":
    mcp.run()
